"""CLI entry point for superintendent disambiguation tool."""

from dotenv import load_dotenv

load_dotenv()

import asyncio
import csv
import logging
import sys
from enum import Enum
from pathlib import Path
from typing import Optional

import click
from openpyxl import Workbook, load_workbook

from content_store import get_store
from disambiguator import MODEL_CONFIGS, create_agent, disambiguate_superintendent
from models import DisambiguationResult, SuperintendentCase, SuperintendentPosition

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)

MAX_RETRIES = 3
INITIAL_BACKOFF = 1.0


class FundsExhausted(Exception):
    """Raised when OpenRouter credits are exhausted."""


def _is_funds_exhausted(error_message: str) -> bool:
    """Detect OpenRouter credit exhaustion from an error message."""
    msg = error_message.lower()
    if "payment required" in msg:
        return True
    if "insufficient" in msg and "credit" in msg:
        return True
    if "out of credits" in msg or "credit balance" in msg:
        return True
    if "openrouter" in msg and "402" in msg:
        return True
    if "openrouter" in msg and "credit" in msg:
        return True
    return False


class FileFormat(Enum):
    """Supported file formats for input/output."""

    XLSX = "xlsx"
    CSV = "csv"


def detect_format(path: Path) -> FileFormat:
    """Detect file format from extension."""
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return FileFormat.XLSX
    elif suffix == ".csv":
        return FileFormat.CSV
    else:
        raise click.BadParameter(f"Unsupported file format: {suffix}. Use .xlsx or .csv")


def _load_cases_xlsx(
    input_path: Path,
    unique_districts_filter: int | None = None,
) -> list[SuperintendentCase]:
    """Load superintendent cases from an Excel file."""
    cases = []
    wb = load_workbook(input_path, read_only=True)
    ws = wb.active

    # Get header row
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {h: i for i, h in enumerate(headers) if h}

    name_idx = header_map.get("name")
    if name_idx is None:
        name_idx = header_map.get("name_clean")
    if name_idx is None:
        name_idx = 0

    unique_districts_idx = header_map.get("unique_districts")
    if unique_districts_filter is not None and unique_districts_idx is None:
        logger.warning(
            "unique_districts filter provided, but column not found in Excel header; "
            "filter will be ignored."
        )

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[name_idx]
        if not name or not str(name).strip():
            continue

        if unique_districts_filter is not None and unique_districts_idx is not None:
            try:
                unique_val = int(row[unique_districts_idx])
            except (ValueError, TypeError):
                logger.warning("Skipping row - invalid unique_districts value")
                continue
            if unique_val != unique_districts_filter:
                continue

        # Parse position 1
        try:
            name_raw1 = str(row[header_map["name_raw1"]]).strip() if header_map.get("name_raw1") and row[header_map["name_raw1"]] else None
            # Handle both "year1" (single year) and "years1" (year range)
            year1_idx = header_map.get("years1") or header_map.get("year1")
            if year1_idx is None:
                raise KeyError("Neither years1 nor year1 found in header")
            year1_val = row[year1_idx]
            # Try to parse as int first, if fails keep as string (for ranges like "2007-2014")
            try:
                year1 = int(year1_val)
            except (ValueError, TypeError):
                year1 = str(year1_val).strip()
            
            pos1 = SuperintendentPosition(
                year=year1,
                state=str(row[header_map["state1"]]).strip(),
                leaid=int(row[header_map["leaid1"]]),
                district_name=str(row[header_map["leaid_name1"]]).strip(),
                name_raw=name_raw1,
            )
        except (KeyError, ValueError, TypeError) as e:
            logger.warning(f"Skipping row - invalid position1 data: {e}")
            continue

        # Parse position 2
        try:
            name_raw2 = str(row[header_map["name_raw2"]]).strip() if header_map.get("name_raw2") and row[header_map["name_raw2"]] else None
            # Handle both "year2" (single year) and "years2" (year range)
            year2_idx = header_map.get("years2") or header_map.get("year2")
            if year2_idx is None:
                raise KeyError("Neither years2 nor year2 found in header")
            year2_val = row[year2_idx]
            # Try to parse as int first, if fails keep as string (for ranges like "2018-2019")
            try:
                year2 = int(year2_val)
            except (ValueError, TypeError):
                year2 = str(year2_val).strip()
            
            pos2 = SuperintendentPosition(
                year=year2,
                state=str(row[header_map["state2"]]).strip(),
                leaid=int(row[header_map["leaid2"]]),
                district_name=str(row[header_map["leaid_name2"]]).strip(),
                name_raw=name_raw2,
            )
        except (KeyError, ValueError, TypeError) as e:
            logger.warning(f"Skipping row - invalid position2 data: {e}")
            continue

        # Parse optional position 3
        pos3 = None
        year3_idx = header_map.get("years3") or header_map.get("year3")
        if year3_idx and row[year3_idx]:
            try:
                name_raw3 = str(row[header_map["name_raw3"]]).strip() if header_map.get("name_raw3") and row[header_map["name_raw3"]] else None
                year3_val = row[year3_idx]
                # Try to parse as int first, if fails keep as string
                try:
                    year3 = int(year3_val)
                except (ValueError, TypeError):
                    year3 = str(year3_val).strip()
                
                pos3 = SuperintendentPosition(
                    year=year3,
                    state=str(row[header_map["state3"]]).strip(),
                    leaid=int(row[header_map["leaid3"]]),
                    district_name=str(row[header_map["leaid_name3"]]).strip(),
                    name_raw=name_raw3,
                )
            except (KeyError, ValueError, TypeError):
                pass  # Position 3 is optional

        case = SuperintendentCase(
            name=str(name).strip(),
            position1=pos1,
            position2=pos2,
            position3=pos3,
        )
        cases.append(case)

    wb.close()
    return cases


def _load_cases_csv(
    input_path: Path,
    unique_districts_filter: int | None = None,
) -> list[SuperintendentCase]:
    """Load superintendent cases from a CSV file."""
    cases = []

    with open(input_path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)

        if unique_districts_filter is not None and "unique_districts" not in reader.fieldnames:
            logger.warning(
                "unique_districts filter provided, but column not found in CSV header; "
                "filter will be ignored."
            )
            unique_districts_filter = None

        for row in reader:
            name = row.get("name") or row.get("name_clean", "")
            if not name or not str(name).strip():
                continue

            if unique_districts_filter is not None:
                try:
                    unique_val = int(row.get("unique_districts", ""))
                except (ValueError, TypeError):
                    logger.warning("Skipping row - invalid unique_districts value")
                    continue
                if unique_val != unique_districts_filter:
                    continue

            # Parse position 1
            try:
                name_raw1 = str(row.get("name_raw1", "")).strip() or None
                # Handle both "year1" (single year) and "years1" (year range)
                year1_val = row.get("years1") or row.get("year1")
                if not year1_val:
                    raise KeyError("Neither years1 nor year1 found")
                # Try to parse as int first, if fails keep as string (for ranges like "2007-2014")
                try:
                    year1 = int(year1_val)
                except (ValueError, TypeError):
                    year1 = str(year1_val).strip()
                
                pos1 = SuperintendentPosition(
                    year=year1,
                    state=str(row["state1"]).strip(),
                    leaid=int(row["leaid1"]),
                    district_name=str(row["leaid_name1"]).strip(),
                    name_raw=name_raw1,
                )
            except (KeyError, ValueError, TypeError) as e:
                logger.warning(f"Skipping row - invalid position1 data: {e}")
                continue

            # Parse position 2
            try:
                name_raw2 = str(row.get("name_raw2", "")).strip() or None
                # Handle both "year2" (single year) and "years2" (year range)
                year2_val = row.get("years2") or row.get("year2")
                if not year2_val:
                    raise KeyError("Neither years2 nor year2 found")
                # Try to parse as int first, if fails keep as string (for ranges like "2018-2019")
                try:
                    year2 = int(year2_val)
                except (ValueError, TypeError):
                    year2 = str(year2_val).strip()
                
                pos2 = SuperintendentPosition(
                    year=year2,
                    state=str(row["state2"]).strip(),
                    leaid=int(row["leaid2"]),
                    district_name=str(row["leaid_name2"]).strip(),
                    name_raw=name_raw2,
                )
            except (KeyError, ValueError, TypeError) as e:
                logger.warning(f"Skipping row - invalid position2 data: {e}")
                continue

            # Parse optional position 3
            pos3 = None
            year3_col = row.get("years3") or row.get("year3")
            if year3_col:
                try:
                    name_raw3 = str(row.get("name_raw3", "")).strip() or None
                    # Try to parse as int first, if fails keep as string
                    try:
                        year3 = int(year3_col)
                    except (ValueError, TypeError):
                        year3 = str(year3_col).strip()
                    
                    pos3 = SuperintendentPosition(
                        year=year3,
                        state=str(row["state3"]).strip(),
                        leaid=int(row["leaid3"]),
                        district_name=str(row["leaid_name3"]).strip(),
                        name_raw=name_raw3,
                    )
                except (KeyError, ValueError, TypeError):
                    pass  # Position 3 is optional

            case = SuperintendentCase(
                name=str(name).strip(),
                position1=pos1,
                position2=pos2,
                position3=pos3,
            )
            cases.append(case)

    return cases


def load_cases(
    input_path: Path,
    unique_districts_filter: int | None = None,
) -> list[SuperintendentCase]:
    """Load superintendent cases from Excel or CSV file."""
    fmt = detect_format(input_path)
    if fmt == FileFormat.XLSX:
        return _load_cases_xlsx(input_path, unique_districts_filter)
    else:
        return _load_cases_csv(input_path, unique_districts_filter)


def _write_results_xlsx(
    output_path: Path,
    cases: list[SuperintendentCase],
    results: list[Optional[DisambiguationResult]],
) -> None:
    """Write results to an Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Disambiguation Results"

    # Write header row
    headers = [
        "name",
        "year1",
        "state1",
        "leaid1",
        "leaid_name1",
        "year2",
        "state2",
        "leaid2",
        "leaid_name2",
        "year3",
        "state3",
        "leaid3",
        "leaid_name3",
        "prediction",
        "confidence",
        "reasoning",
        "evidence_urls",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Write data rows
    for row_num, (case, (result, error_msg)) in enumerate(zip(cases, results), 2):
        ws.cell(row=row_num, column=1, value=case.name)
        ws.cell(row=row_num, column=2, value=case.position1.year)
        ws.cell(row=row_num, column=3, value=case.position1.state)
        ws.cell(row=row_num, column=4, value=case.position1.leaid)
        ws.cell(row=row_num, column=5, value=case.position1.district_name)
        ws.cell(row=row_num, column=6, value=case.position2.year)
        ws.cell(row=row_num, column=7, value=case.position2.state)
        ws.cell(row=row_num, column=8, value=case.position2.leaid)
        ws.cell(row=row_num, column=9, value=case.position2.district_name)

        if case.position3:
            ws.cell(row=row_num, column=10, value=case.position3.year)
            ws.cell(row=row_num, column=11, value=case.position3.state)
            ws.cell(row=row_num, column=12, value=case.position3.leaid)
            ws.cell(row=row_num, column=13, value=case.position3.district_name)

        if result:
            ws.cell(row=row_num, column=14, value=result.prediction)
            ws.cell(row=row_num, column=15, value=result.confidence)
            ws.cell(row=row_num, column=16, value=result.reasoning)
            ws.cell(row=row_num, column=17, value=", ".join(result.evidence_urls))
        else:
            ws.cell(row=row_num, column=14, value="error")
            ws.cell(row=row_num, column=15, value="")
            ws.cell(row=row_num, column=16, value=error_msg if error_msg else "Processing failed")
            ws.cell(row=row_num, column=17, value="")

    wb.save(output_path)


# Shared headers for both xlsx and csv
RESULT_HEADERS = [
    "name",
    "year1",
    "state1",
    "leaid1",
    "leaid_name1",
    "year2",
    "state2",
    "leaid2",
    "leaid_name2",
    "year3",
    "state3",
    "leaid3",
    "leaid_name3",
    "prediction",
    "confidence",
    "reasoning",
    "evidence_urls",
]


def _case_to_row_dict(
    case: SuperintendentCase, result: Optional[DisambiguationResult], error_msg: str = ""
) -> dict:
    """Convert a case and result to a row dictionary."""
    row = {
        "name": case.name,
        "year1": case.position1.year,
        "state1": case.position1.state,
        "leaid1": case.position1.leaid,
        "leaid_name1": case.position1.district_name,
        "year2": case.position2.year,
        "state2": case.position2.state,
        "leaid2": case.position2.leaid,
        "leaid_name2": case.position2.district_name,
        "year3": case.position3.year if case.position3 else "",
        "state3": case.position3.state if case.position3 else "",
        "leaid3": case.position3.leaid if case.position3 else "",
        "leaid_name3": case.position3.district_name if case.position3 else "",
    }

    if result:
        row["prediction"] = result.prediction
        row["confidence"] = result.confidence
        row["reasoning"] = result.reasoning
        row["evidence_urls"] = ", ".join(result.evidence_urls)
    else:
        row["prediction"] = "error"
        row["confidence"] = ""
        row["reasoning"] = error_msg if error_msg else "Processing failed"
        row["evidence_urls"] = ""

    return row


def _write_results_csv(
    output_path: Path,
    cases: list[SuperintendentCase],
    results: list[tuple[Optional[DisambiguationResult], str]],
) -> None:
    """Write results to a CSV file."""
    with open(output_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=RESULT_HEADERS)
        writer.writeheader()

        for case, (result, error_msg) in zip(cases, results):
            row = _case_to_row_dict(case, result, error_msg)
            writer.writerow(row)


def write_results(
    output_path: Path,
    cases: list[SuperintendentCase],
    results: list[tuple[Optional[DisambiguationResult], str]],
) -> None:
    """Write results to Excel or CSV file."""
    fmt = detect_format(output_path)
    if fmt == FileFormat.XLSX:
        _write_results_xlsx(output_path, cases, results)
    else:
        _write_results_csv(output_path, cases, results)


def _init_output_file(output_path: Path, fmt: FileFormat) -> None:
    """Initialize output file with header only. Overwrites if exists."""
    if fmt == FileFormat.CSV:
        with open(output_path, "w", encoding="utf-8", newline="") as f:
            csv.DictWriter(f, fieldnames=RESULT_HEADERS).writeheader()
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Disambiguation Results"
        ws.append(RESULT_HEADERS)
        wb.save(output_path)


def append_result(
    output_path: Path,
    case: SuperintendentCase,
    result: Optional[DisambiguationResult],
    error_msg: str,
    fmt: FileFormat,
) -> None:
    """Append a single result row to output file."""
    row = _case_to_row_dict(case, result, error_msg)
    row_values = [row[h] for h in RESULT_HEADERS]
    if fmt == FileFormat.CSV:
        with open(output_path, "a", encoding="utf-8", newline="") as f:
            csv.DictWriter(f, fieldnames=RESULT_HEADERS).writerow(row)
    else:
        wb = load_workbook(output_path)
        ws = wb.active
        ws.append(row_values)
        wb.save(output_path)


async def process_case_with_retry(
    agent,
    case: SuperintendentCase,
) -> tuple[Optional[DisambiguationResult], str]:
    """Process a single case with exponential backoff retry logic.

    Returns tuple of (result, error_message).
    If successful, error_message is empty string.
    """
    backoff = INITIAL_BACKOFF
    last_error = ""

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            result = await disambiguate_superintendent(agent, case)
            return result, ""
        except Exception as e:
            last_error = str(e)
            error_type = type(e).__name__
            logger.warning(f"Attempt {attempt}/{MAX_RETRIES} failed for {case.name}: [{error_type}] {e}")
            if "exceeds this API key's set usage limit" in last_error.lower():
                logger.error("Tavily quota exceeded; skipping remaining retries for this case")
                return None, f"{error_type}: {last_error[:100]}"
            if _is_funds_exhausted(last_error):
                logger.error("OpenRouter credits exhausted; stopping run")
                raise FundsExhausted(last_error)
            if attempt < MAX_RETRIES:
                logger.info(f"Retrying in {backoff:.1f} seconds...")
                await asyncio.sleep(backoff)
                backoff *= 2
            else:
                logger.error(f"All retries exhausted for {case.name}: {error_type}")
                return None, f"{error_type}: {last_error[:100]}"


async def process_all_cases(
    input_path: Path,
    output_path: Path,
    model_name: str = "glm-4.6",
    unique_districts_filter: int | None = None,
) -> None:
    """Process all cases from input file and write results. Writes incrementally after each case."""
    cases = load_cases(input_path, unique_districts_filter)
    logger.info(f"Loaded {len(cases)} cases from {input_path}")

    if not cases:
        logger.error("No cases found in input file")
        return

    fmt = detect_format(output_path)
    _init_output_file(output_path, fmt)

    agent = create_agent(model_name)
    logger.info(f"Using model: {model_name}")
    results: list[tuple[Optional[DisambiguationResult], str]] = []

    funds_exhausted_msg = ""

    async with agent:
        for idx, case in enumerate(cases):
            districts = f"{case.position1.district_name} -> {case.position2.district_name}"
            if case.position3:
                districts += f" -> {case.position3.district_name}"
            logger.info(f"Processing {idx + 1}/{len(cases)}: {case.name} ({districts})")

            try:
                result, error_msg = await process_case_with_retry(agent, case)
                results.append((result, error_msg))
            except FundsExhausted as e:
                funds_exhausted_msg = f"OpenRouter credits exhausted: {str(e)[:120]}"
                logger.error(funds_exhausted_msg)
                results.append((None, funds_exhausted_msg))
                append_result(output_path, case, None, funds_exhausted_msg, fmt)

                for rem_case in cases[idx + 1 :]:
                    results.append((None, funds_exhausted_msg))
                    append_result(output_path, rem_case, None, funds_exhausted_msg, fmt)
                get_store().clear()
                break

            append_result(output_path, case, result, error_msg, fmt)

            if result:
                # Truncate reasoning for log display
                reasoning_text = result.reasoning
                logger.info(f"DEBUG Full reasoning ({len(reasoning_text)} chars): {reasoning_text!r}")
                reasoning_display = reasoning_text[:200] + "..." if len(reasoning_text) > 200 else reasoning_text
                logger.info(
                    f"Result: {case.name}, {districts}, "
                    f"{result.prediction.upper()} ({result.confidence}). {reasoning_display}"
                )
            elif error_msg:
                logger.error(f"Failed with error: {error_msg}")

            # Clear content store between cases to prevent accumulation
            get_store().clear()

        if funds_exhausted_msg:
            logger.error("Stopped early due to exhausted credits")

    logger.info(f"Results written to {output_path}")


@click.group()
def cli() -> None:
    """Superintendent disambiguation tool."""
    pass


@cli.command()
@click.argument("input_file", type=click.Path(exists=True, path_type=Path))
@click.argument("output_file", type=click.Path(path_type=Path))
@click.option("-v", "--verbose", is_flag=True, help="Enable debug logging")
@click.option(
    "--unique-districts",
    "unique_districts",
    type=int,
    default=None,
    help="Only process rows where unique_districts matches this value",
)
@click.option(
    "-m",
    "--model",
    type=click.Choice(MODEL_CONFIGS),
    default="z-ai/glm-4.6",
    help="LLM model to use (via OpenRouter)",
)
def process(
    input_file: Path,
    output_file: Path,
    verbose: bool,
    unique_districts: int | None,
    model: str,
) -> None:
    """Process superintendent cases from file.

    INPUT_FILE: Path to input file with superintendent cases (.xlsx or .csv)
    OUTPUT_FILE: Path to output file for results (.xlsx or .csv)
    """
    if verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    try:
        asyncio.run(process_all_cases(input_file, output_file, model, unique_districts))
    except KeyboardInterrupt:
        logger.info("Interrupted by user")
        sys.exit(1)
    except Exception as e:
        logger.error(f"Fatal error: {e}")
        sys.exit(1)
    else:
        logger.info("Done.")


async def run_single_test(
    name: str,
    district1: str,
    state1: str,
    year1: int,
    district2: str,
    state2: str,
    year2: int,
    model_name: str,
) -> None:
    """Run disambiguation on a single test case."""
    case = SuperintendentCase(
        name=name,
        position1=SuperintendentPosition(
            year=year1,
            state=state1,
            leaid=0,  # Not needed for testing
            district_name=district1,
        ),
        position2=SuperintendentPosition(
            year=year2,
            state=state2,
            leaid=0,
            district_name=district2,
        ),
    )

    agent = create_agent(model_name)
    logger.info(f"Testing: {name}")
    logger.info(f"  Position 1: {district1} ({state1}), {year1}")
    logger.info(f"  Position 2: {district2} ({state2}), {year2}")
    logger.info(f"  Model: {model_name}")
    logger.info("-" * 60)

    async with agent:
        result, error_msg = await process_case_with_retry(agent, case)

    if result:
        logger.info(f"\n{'='*60}")
        logger.info(f"RESULT: {result.prediction.upper()} (confidence: {result.confidence})")
        logger.info(f"{'='*60}")
        logger.info(f"\nREASONING:\n{result.reasoning}")
        if result.evidence_urls:
            logger.info("\nEVIDENCE URLs:")
            for url in result.evidence_urls:
                logger.info(f"  - {url}")
    else:
        logger.error("Test failed - no result returned")
        if error_msg:
            logger.error("Error: %s", error_msg)


@cli.command()
@click.option("--name", required=True, help="Superintendent name")
@click.option("--district1", required=True, help="First district name")
@click.option("--state1", required=True, help="First state (2-letter code)")
@click.option("--year1", required=True, type=int, help="First position year")
@click.option("--district2", required=True, help="Second district name")
@click.option("--state2", required=True, help="Second state (2-letter code)")
@click.option("--year2", required=True, type=int, help="Second position year")
@click.option("-v", "--verbose", is_flag=True, help="Enable debug logging")
@click.option(
    "-m",
    "--model",
    type=click.Choice(MODEL_CONFIGS),
    default="z-ai/glm-4.6",
    help="LLM model to use (via OpenRouter)",
)
def test(
    name: str,
    district1: str,
    state1: str,
    year1: int,
    district2: str,
    state2: str,
    year2: int,
    verbose: bool,
    model: str,
) -> None:
    """Test disambiguation on a single case (no Excel required).

    Example:
        uv run python main.py test --name "Emin Cavusoglu" \\
            --district1 "Lisa Academy North Charter" --state1 AR --year1 2008 \\
            --district2 "School of Science and Technology Discovery" --state2 TX --year2 2012
    """
    if verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    try:
        asyncio.run(
            run_single_test(name, district1, state1, year1, district2, state2, year2, model)
        )
    except KeyboardInterrupt:
        logger.info("Interrupted by user")
        sys.exit(1)
    except Exception as e:
        logger.error(f"Fatal error: {e}")
        sys.exit(1)


def main() -> None:
    """Entry point."""
    cli()


if __name__ == "__main__":
    main()
